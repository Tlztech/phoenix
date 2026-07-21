from __future__ import annotations

import hashlib
import logging
import json
import os
import re
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook

from .checkpoint import Checkpoint
from .constants import DEFAULT_CATEGORY_URL
from .excel import ExcelOutput
from .fetcher import FetchConfig, PatagoniaFetcher
from .models import ProductPageData, ProductRow, dedupe_rows
from .parser import (
    canonical_product_url,
    discover_product_urls,
    extract_image_urls,
    find_hero_image,
    image_basename,
    images_for_color,
    model_from_url,
    order_gallery,
    parse_product_page,
    format_yen,
)
from .generic_parser import parse_generic_product_page

LOGGER = logging.getLogger(__name__)


@dataclass(slots=True)
class ScraperConfig:
    category_urls: list[str] = field(default_factory=lambda: [DEFAULT_CATEGORY_URL])
    max_pages: int = 20
    max_products: int | None = None
    url_file: str | None = None
    output_dir: Path = Path("output")
    resume: bool = True
    resume_max_age_hours: float = 24.0  # ignore (and refresh) checkpoints older than this
    flush_every: int = 5
    stock_cap: int = 5  # cap quantity at this (the dropdown max); 0 = report real ATS
    fetch: FetchConfig = field(default_factory=FetchConfig)


@dataclass(slots=True)
class ScrapeResult:
    rows: list[ProductRow]
    total: int
    done: int
    complete: bool
    partial_path: Path
    checkpoint: Checkpoint
    discovery_failed: bool = False
    failed_categories: list[str] = field(default_factory=list)


def _read_url_file(path: str | Path) -> list[str]:
    """Collect product URLs from any column of a workbook, de-duplicated.

    The URLs are used verbatim (only canonicalised); the live site resolves the
    ``/product/<slug>/<id>.html`` path itself, so we must not rewrite it.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            value = str(cell).strip() if cell else ""
            if "/product/" in value:
                values.append(canonical_product_url(value))
    return list(dict.fromkeys(values))


def _variant_url(url: str, model: str, color: str, upc: str = "") -> str:
    if not color:
        return canonical_product_url(url)
    parts = urlsplit(url)
    if upc.isdigit() and len(upc) >= 10:
        segments = parts.path.rstrip("/").split("/")
        if segments:
            segments[-1] = f"{upc}.html"
            parts = parts._replace(path="/".join(segments))
            return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))
    query = [(key, value) for key, value in parse_qsl(parts.query) if not key.lower().startswith("dwvar_")]
    query.append((f"dwvar_{model}_color", color))
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), ""))


class PatagoniaScraper:
    def __init__(self, config: ScraperConfig | None = None) -> None:
        self.config = config or ScraperConfig()
        self.client = PatagoniaFetcher(self.config.fetch)
        self._flush_lock = threading.Lock()
        self._failed_categories: list[str] = []

    def _job_key(self) -> str:
        """Stable id for the current job so its checkpoint/partial files are reused."""
        if self.config.url_file:
            basis = f"urlfile:{Path(self.config.url_file).resolve()}"
        else:
            basis = "cats:" + "|".join(sorted(self.config.category_urls))
        return hashlib.md5(basis.encode("utf-8")).hexdigest()[:8]

    def _flush_partial(self, checkpoint: Checkpoint, path: Path) -> None:
        with self._flush_lock:
            rows, _ = dedupe_rows(checkpoint.all_rows())
            try:
                writer = ExcelOutput(path)
                writer.append(rows)
                writer.save()
            except Exception as exc:  # a locked/open file should not kill the crawl
                LOGGER.warning("Partial save to %s failed: %s", path, exc)

    def _discover_category(self, base: str) -> list[str]:
        """Fetch one category page and return its product URLs, retrying past blocks.

        If Akamai serves a 404 (empty grid), we escalate: clear cookies and fully
        re-warm the session, then try again a couple more times.
        """
        items: list[str] = []
        for attempt in range(1, 4):
            LOGGER.info("Fetching category page: %s (第%s次)", base, attempt)
            page = self.client.fetch(base, page_action=self._expand_catalogue)
            items = discover_product_urls(page, base)
            if items:
                return items
            LOGGER.warning(
                "类别 %s 第%s次未发现商品（status %s，可能被 Akamai 拦截）",
                base,
                attempt,
                getattr(page, "status", "?"),
            )
            if attempt < 3:
                self.client.force_rewarm(clear_cookies=True)
        return items

    def product_urls(self) -> list[str]:
        self._failed_categories = []
        if self.config.url_file:
            values = _read_url_file(self.config.url_file)
            return values[: self.config.max_products] if self.config.max_products else values
        found: list[str] = []
        seen: set[str] = set()
        for base in self.config.category_urls:
            items = self._discover_category(base)
            if not items:
                self._failed_categories.append(base)
                LOGGER.error("类别页最终未发现任何商品（放弃该类别）: %s", base)
                continue
            for item in items:
                item = canonical_product_url(item, base)
                if item not in seen:
                    seen.add(item)
                    found.append(item)
        LOGGER.info(
            "Discovered %s product URLs across %s categories (%s 个类别失败)",
            len(found),
            len(self.config.category_urls),
            len(self._failed_categories),
        )
        if self.config.max_products:
            return found[: self.config.max_products]
        return found

    def _expand_catalogue(self, browser_page: object) -> None:
        """Scroll and click the "load more" control until the grid stops growing.

        The womens grid lazy-loads product cards on scroll and exposes a
        ``さらに見る`` / ``もっと見る`` button. We repeat scroll + click until the
        product-link count is stable across a couple of rounds (or we hit the
        ``max_pages`` safety bound).
        """
        rounds = max(1, self.config.max_pages) * 3
        previous = -1
        stable = 0
        for _ in range(rounds):
            try:
                browser_page.mouse.wheel(0, 5_000)
                browser_page.wait_for_timeout(1_100)
                for label in ("さらに見る", "もっと見る"):
                    button = browser_page.get_by_text(label, exact=True)
                    if button.count() > 0:
                        try:
                            button.first.click(timeout=2_000)
                            browser_page.wait_for_timeout(1_200)
                        except Exception:
                            pass
                count = browser_page.eval_on_selector_all(
                    'a[href*="/product/"]', "els => els.length"
                )
            except Exception:
                break
            if count <= previous:
                stable += 1
                if stable >= 3:
                    break
            else:
                stable = 0
            previous = count

    def scrape(self) -> ScrapeResult:
        output_dir = Path(self.config.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        job = self._job_key()
        checkpoint_path = output_dir / f".checkpoint_{job}.jsonl"
        partial_path = output_dir / f"patagonia_output_{job}_part.xlsx"
        # A checkpoint only exists when a previous run was interrupted. If it is
        # stale (e.g. a periodic re-run days later), discard it so we always fetch
        # fresh data instead of resuming and merging outdated rows.
        if self.config.resume and checkpoint_path.exists():
            age_hours = (time.time() - checkpoint_path.stat().st_mtime) / 3600.0
            if age_hours > self.config.resume_max_age_hours:
                LOGGER.info(
                    "Checkpoint is %.1fh old (> %.0fh); discarding it and scraping fresh",
                    age_hours,
                    self.config.resume_max_age_hours,
                )
                for stale in (
                    checkpoint_path,
                    partial_path,
                    partial_path.with_name(partial_path.stem + "_summary.txt"),
                ):
                    try:
                        stale.unlink()
                    except OSError:
                        pass
        checkpoint = Checkpoint(checkpoint_path, enabled=self.config.resume)
        flush_every = max(1, self.config.flush_every)
        urls: list[str] = []
        try:
            urls = self.product_urls()
            total = len(urls)
            if total == 0:
                # Nothing discovered — almost always an Akamai block on the category
                # pages (or an empty URL file). Do not touch the checkpoint or write a
                # misleading partial; let the caller report and the user retry.
                LOGGER.error(
                    "未发现任何商品 URL；类别页很可能被 Akamai 拦截。断点(%s 个商品)已保留。",
                    checkpoint.done_count,
                )
                return ScrapeResult(
                    checkpoint.all_rows(), 0, checkpoint.done_count, False, partial_path, checkpoint,
                    discovery_failed=True, failed_categories=list(self._failed_categories),
                )
            pending = [url for url in urls if not checkpoint.is_done(canonical_product_url(url))]
            if checkpoint.done_count:
                LOGGER.info(
                    "Resuming: %s/%s already done, %s pending", checkpoint.done_count, total, len(pending)
                )

            # Test-only hook: PATAGONIA_SIM_STOP_AFTER=N makes a FRESH run abort after
            # N products (as if rate-limited), so the cooldown/resume flow can be
            # rehearsed quickly. Ignored on resume (checkpoint already has products).
            sim_stop = int(os.environ.get("PATAGONIA_SIM_STOP_AFTER", "0") or 0)
            fresh_run = checkpoint.done_count == 0

            def handler(fetch, url: str) -> list[ProductRow]:
                key = canonical_product_url(url)
                page = fetch(url)
                self._ensure_product_response(page, url)
                try:
                    product = parse_product_page(page, url)
                except (ValueError, json.JSONDecodeError, KeyError, TypeError, AttributeError):
                    LOGGER.warning("Product schema variant not found; using generic parser: %s", url)
                    product = parse_generic_product_page(page, url)
                rows = self._rows_for_product(product, page, url, fetch)
                done = checkpoint.add(key, rows)  # flushed to disk immediately
                LOGGER.info("Done %s/%s: %s", done, total, url)
                if done % flush_every == 0:
                    self._flush_partial(checkpoint, partial_path)
                if sim_stop and fresh_run and done >= sim_stop:
                    LOGGER.warning("[模拟] 已抓 %s 个，模拟被限流 -> 中止本轮", done)
                    self.client._abort.set()
                return rows

            self.client.run_pool(pending, handler)
        finally:
            self.client.close()

        self._flush_partial(checkpoint, partial_path)
        rows = checkpoint.all_rows()
        remaining = [url for url in urls if not checkpoint.is_done(canonical_product_url(url))]
        # A category that returned nothing (blocked) means the discovered set is
        # incomplete, so we must not declare the job finished even if every
        # discovered URL was scraped.
        complete = len(urls) > 0 and len(remaining) == 0 and not self._failed_categories
        LOGGER.info("Scraped %s products (%s rows); complete=%s", checkpoint.done_count, len(rows), complete)
        return ScrapeResult(
            rows, len(urls), checkpoint.done_count, complete, partial_path, checkpoint,
            failed_categories=list(self._failed_categories),
        )

    @staticmethod
    def _ensure_product_response(page: object, url: str) -> None:
        status = getattr(page, "status", None)
        if status is None:
            status = getattr(page, "status_code", None)
        if status is not None and int(status) >= 400:
            raise RuntimeError(f"HTTP {status} for {url} (Akamai block or missing product)")
        body = ""
        if hasattr(page, "get_all_text"):
            try:
                body = page.get_all_text()
            except Exception:
                body = ""
        if not body:
            body = getattr(page, "text", "") or ""
        if isinstance(body, bytes):
            body = body.decode("utf-8", "ignore")
        lowered = str(body).lower()
        if any(marker in lowered for marker in ("page not found", "error 404", "access denied")):
            raise RuntimeError(f"Blocked or missing product page: {url}")

    def _color_images(self, color, model: str, page: object, fetch) -> tuple[str, list[str]]:
        """Return ``(main_image, other_images)`` for one colour.

        The hero shot (``{model}_{color}.jpg``) is on the product page; the detail
        gallery is loaded per colour from Patagonia's ``Product-AltAssets``
        endpoint, so we fetch it in the same warm session.
        """
        all_images = extract_image_urls(page)
        main = find_hero_image(all_images, model, color.code) or color.main_image or (
            all_images[0] if all_images else ""
        )
        gallery: list[str] = []
        if color.alt_assets_url:
            try:
                alt_page = fetch(color.alt_assets_url, scroll=False)
                if getattr(alt_page, "status", 0) < 400:
                    gallery = extract_image_urls(alt_page)
            except Exception:
                LOGGER.debug("Alt-assets fetch failed for %s/%s", model, color.code)
        if not gallery:
            gallery = images_for_color(all_images, model, color.code, main)
        main_base = image_basename(main).lower()
        gallery = [u for u in gallery if image_basename(u).lower() != main_base]
        return main, order_gallery(gallery)

    def _variant_data(self, model: str, color: str, size: str, endpoint: str, fetch) -> dict | None:
        """Per-variant availability + price from SFCC's ``Product-Variation`` endpoint.

        Returns ``{"ats", "status", "list", "sales"}`` or ``None`` if unavailable.
        ``list`` is the original price (msrp), ``sales`` the current price.
        """
        if not (endpoint and color):
            return None
        url = f"{endpoint}?pid={model}&dwvar_{model}_color={color}"
        if size:
            url += f"&dwvar_{model}_size={size}"
        url += "&quantity=1"
        try:
            resp = fetch(url, scroll=False)
            if getattr(resp, "status", 0) >= 400:
                return None
            data = json.loads(resp.get_all_text())
        except Exception:
            return None
        prod = (data or {}).get("product") or {}
        availability = prod.get("availability") or {}
        price = prod.get("price") or {}

        def _value(node: object) -> object:
            return node.get("value") if isinstance(node, dict) else None

        raw = availability.get("ATS")
        if raw is None:
            raw = availability.get("stockLevel")
        try:
            ats = int(float(raw)) if raw is not None else None
        except (TypeError, ValueError):
            ats = None
        return {
            "ats": ats,
            "status": str(availability.get("stockStatus") or "").strip(),
            "list": _value(price.get("list")),
            "sales": _value(price.get("sales")),
        }

    def _rows_for_product(self, product: ProductPageData, page: object, page_url: str, fetch) -> list[ProductRow]:
        rows: list[ProductRow] = []
        endpoint = product.variation_endpoint
        for color in product.colors:
            main, other = self._color_images(color, product.model, page, fetch)
            sizes = color.sizes or [""]

            # Query the variation endpoint: the first size (for price, even if OOS)
            # plus every in-stock size (for its real available quantity).
            data_by_size: dict[str, dict] = {}
            if endpoint:
                first = self._variant_data(product.model, color.code, sizes[0], endpoint, fetch)
                if first:
                    data_by_size[sizes[0]] = first
                for size in sizes:
                    if size in data_by_size:
                        continue
                    if size in color.in_stock_sizes and size not in color.out_of_stock_upcs:
                        info = self._variant_data(product.model, color.code, size, endpoint, fetch)
                        if info:
                            data_by_size[size] = info

            # msrp = original (list) price; discounted only when it differs.
            color_list = color_sales = None
            for size in sizes:
                info = data_by_size.get(size)
                if info and info["list"] is not None:
                    color_list, color_sales = info["list"], info["sales"]
                    break
            if color_list is not None:
                msrp = format_yen(color_list)
                discounted = format_yen(color_sales) if (color_sales is not None and color_sales != color_list) else ""
            else:  # endpoint gave no price — fall back to the JSON-LD values
                msrp = format_yen(color.msrp if color.msrp not in (None, "") else color.offer_price)
                discounted = format_yen(color.discounted_price)

            for size in sizes:
                # A one-size "ALL" variant is written with an empty size and a
                # ``model-color`` sku (no "-ALL" suffix).
                is_all = size.strip().upper() == "ALL"
                out_size = "" if is_all else size
                if out_size:
                    sku = color.sku_by_size.get(size, "") or f"{product.model}-{color.code}-{out_size}"
                else:
                    sku = f"{product.model}-{color.code}".strip("-")
                # upc is only used to build the variant URL; the output column stays empty.
                upc = color.upc_by_size.get(size, "") or color.out_of_stock_upcs.get(size, "")
                in_stock = size in color.in_stock_sizes and size not in color.out_of_stock_upcs
                quantity = 0
                status = "在庫なし"
                if in_stock:
                    info = data_by_size.get(size)
                    if info and info["ats"] is not None:
                        ats = info["ats"]
                        if self.config.stock_cap and self.config.stock_cap > 0:
                            ats = min(ats, self.config.stock_cap)
                        quantity = max(0, ats)  # oversold variants report a negative ATS → treat as 0
                        status = "在庫あり" if quantity > 0 else "在庫なし"
                    else:  # endpoint unavailable — fall back to a binary in-stock flag
                        quantity, status = 1, "在庫あり"
                row = ProductRow(
                    title=product.title,
                    model=product.model or model_from_url(page_url),
                    sku=sku,
                    upc_barcode_ean="",
                    brand=product.brand or "patagonia",
                    size=out_size,
                    url=_variant_url(color.variant_urls.get(size, color.url or page_url), product.model, color.code, upc),
                    color=color.code,
                    msrp=msrp,
                    discounted_price=discounted,
                    product_main_image=main,
                    product_other_image="\n".join(other),
                    dimension=product.dimension,
                    description=product.description,
                    product_spec=color.product_spec,
                    material=product.material,
                    weight=product.weight,
                    stock_status=status,
                    quantity=str(quantity),
                )
                rows.append(row)
        return rows
