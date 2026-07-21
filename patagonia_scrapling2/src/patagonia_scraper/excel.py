from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import OUTPUT_COLUMNS, ProductRow


def timestamped_path(output_dir: str | Path = ".") -> Path:
    """Return the required local-time output filename."""
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return Path(output_dir) / f"patagonia_output_{stamp}.xlsx"


class ExcelOutput:
    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Sheet1"
        self.sheet.append(OUTPUT_COLUMNS)
        self._style_header()

    def _style_header(self) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in self.sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.freeze_panes = "A2"
        self.sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"

    def append(self, rows: list[ProductRow]) -> None:
        for row in rows:
            self.sheet.append(row.values())

    def save(self) -> None:
        widths = {
            "A": 42,
            "B": 12,
            "C": 28,
            "D": 18,
            "E": 14,
            "F": 12,
            "G": 70,
            "H": 14,
            "I": 16,
            "J": 18,
            "K": 70,
            "L": 70,
            "M": 55,
            "N": 70,
            "O": 40,
            "P": 70,
            "Q": 16,
            "R": 14,
            "S": 10,
        }
        for column, width in widths.items():
            self.sheet.column_dimensions[column].width = width
        self.sheet.auto_filter.ref = self.sheet.dimensions
        for row in self.sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        self.workbook.save(self.path)
