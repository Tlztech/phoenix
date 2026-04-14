from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path

from deep_translator import GoogleTranslator
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
TITLE_COL_NAME = "Title"
TITLE_CHN_COL_NAME = "TitleChn"
CACHE_FILE = SCRIPT_DIR / "title_translation_cache.json"

KANA_RE = re.compile(r"[\u3040-\u30ff\u31f0-\u31ff]")
CHINESE_RE = re.compile(r"[\u4e00-\u9fff]")

GENDER_PATTERNS = [
    (re.compile(r"(?i)\bmen'?s\b"), "男款"),
    (re.compile(r"(?i)\bwomen'?s\b"), "女款"),
    (re.compile(r"(?i)\bkid'?s\b"), "儿童款"),
    (re.compile(r"(?i)\bkids\b"), "儿童款"),
    (re.compile(r"\bメンズ\b"), "男款"),
    (re.compile(r"\bウィメンズ\b"), "女款"),
    (re.compile(r"\bレディース\b"), "女款"),
    (re.compile(r"\bキッズ\b"), "儿童款"),
]

PROTECTED_TERMS = {
    "montbell": "Montbell",
    "wic.": "__WIC__",
    "geoline": "__GEOLINE__",
    "superior down": "__SUPERIOR_DOWN__",
    "trail action": "__TRAIL_ACTION__",
    "zeo-line": "__ZEO_LINE__",
    "dry-tec": "__DRY_TEC__",
    "dry tec": "__DRY_TEC__",
    "ex light": "__EX_LIGHT__",
}

DIRECT_MAP = {
    "ステンレスメッシュ キャップシェード": "不锈钢网罩",
    "ステンレスメッシュ ハット": "不锈钢网帽",
    "ステンレスメッシュ ハットシェード": "不锈钢网帽遮阳帘",
    "Montbell サンブロック ハットスクリーン": "Montbell 防晒帽屏风",
    "クール パーカ 男款": "男款凉感连帽外套",
    "クール パーカ 女款": "女款凉感连帽外套",
    "トレール アームカバー ライト 男款": "轻便男款步履护臂",
    "トレール レッグカバー ライト 女款": "轻便女款步履腿套",
    "スペリオダウンジャケット 男款": "男款 Superior 轻量羽绒夹克",
    "スペリオダウンジャケット 女款": "女款 Superior 轻量羽绒夹克",
    "EXライトダウン": "EX超轻羽绒外套",
}

STYLE_REPLACEMENTS = [
    (re.compile(r"__WIC__"), "WIC."),
    (re.compile(r"__GEOLINE__"), "Geoline"),
    (re.compile(r"__SUPERIOR_DOWN__"), "Superior"),
    (re.compile(r"__TRAIL_ACTION__"), "Trail Action"),
    (re.compile(r"__ZEO_LINE__"), "ZEO-LINE"),
    (re.compile(r"__DRY_TEC__"), "Dry-Tec"),
    (re.compile(r"__EX_LIGHT__"), "EX超轻"),
    (re.compile(r"\b男士\b"), "男款"),
    (re.compile(r"\b女士\b"), "女款"),
    (re.compile(r"\b儿童\b"), "儿童款"),
    (re.compile(r"\b男袜\b"), "男款"),
    (re.compile(r"\b女袜\b"), "女款"),
    (re.compile(r"\b派克大衣\b"), "连帽外套"),
    (re.compile(r"\b派克外套\b"), "连帽外套"),
    (re.compile(r"\b风衣\b"), "外套"),
    (re.compile(r"\b帽檐护帘\b"), "帽屏风"),
    (re.compile(r"\b帽屏\b"), "帽屏风"),
    (re.compile(r"\b护臂套\b"), "护臂"),
    (re.compile(r"\b臂套\b"), "护臂"),
    (re.compile(r"\b袖套\b"), "护臂"),
    (re.compile(r"\b凉爽\b"), "凉感"),
    (re.compile(r"\b不锈钢网眼帽\b"), "不锈钢网帽"),
    (re.compile(r"\b不锈钢网眼\b"), "不锈钢网"),
    (re.compile(r"\b越野 护臂 轻量\b"), "轻便步履护臂"),
    (re.compile(r"\b越野 腿套 轻量\b"), "轻便步履腿套"),
    (re.compile(r"\b美利奴羊毛高山袜 男款\b"), "男款美利奴羊毛高山袜"),
    (re.compile(r"\b美利奴羊毛高山袜 女款\b"), "女款美利奴羊毛高山袜"),
]


def contains_kana(text: str) -> bool:
    return bool(KANA_RE.search(text))


def contains_chinese(text: str) -> bool:
    return bool(CHINESE_RE.search(text))


def should_skip_translation(text: str) -> bool:
    return contains_chinese(text)


def normalize_gender(text: str) -> str:
    result = text
    for pattern, replacement in GENDER_PATTERNS:
        result = pattern.sub(replacement, result)
    return re.sub(r"\s+", " ", result).strip()


def protect_terms(text: str) -> str:
    result = text
    for source, target in PROTECTED_TERMS.items():
        result = re.sub(re.escape(source), target, result, flags=re.IGNORECASE)
    return result


def normalize_before_translation(text: str) -> str:
    normalized = normalize_gender(text.strip())
    normalized = protect_terms(normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def cleanup_after_translation(text: str) -> str:
    cleaned = text.strip()
    for pattern, replacement in STYLE_REPLACEMENTS:
        cleaned = pattern.sub(replacement, cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"\s+([,./])", r"\1", cleaned)
    cleaned = re.sub(r"^(.*?)[\s/-]+(男款|女款|儿童款)$", r"\2 \1", cleaned)
    cleaned = re.sub(r"^(Montbell)\s+(男款|女款|儿童款)\s+", r"\1 \2 ", cleaned)
    return cleaned.strip()


def load_cache(cache_path: Path) -> dict[str, str]:
    if not cache_path.exists():
        return {}
    return json.loads(cache_path.read_text(encoding="utf-8"))


def save_cache(cache_path: Path, cache: dict[str, str]) -> None:
    cache_path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def translate_with_retry(
    translator: GoogleTranslator,
    text: str,
    cache: dict[str, str],
    retries: int = 3,
) -> str:
    if text in cache:
        return cache[text]

    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            translated = translator.translate(text) or text
            cache[text] = translated
            return translated
        except Exception as exc:  # pragma: no cover
            last_error = exc
            time.sleep(1.5 * (attempt + 1))

    raise RuntimeError(f"Failed to translate: {text}") from last_error


def translate_title(title: str, translator: GoogleTranslator, cache: dict[str, str]) -> str:
    source = str(title).strip()
    if not source:
        return source
    if should_skip_translation(source):
        return source

    normalized = normalize_before_translation(source)
    if normalized in DIRECT_MAP:
        return DIRECT_MAP[normalized]
    if not contains_kana(normalized):
        return cleanup_after_translation(normalized)

    translated = translate_with_retry(translator, normalized, cache)
    return cleanup_after_translation(translated)


def resolve_input_path(raw_path: str | None) -> Path:
    if raw_path:
        return Path(raw_path).expanduser()

    download_dir = Path.home() / "Downloads"
    preferred = download_dir / "MONTBELL_普通款总表.xlsx"
    if preferred.exists():
        return preferred

    matches = sorted(download_dir.glob("MONTBELL_*.xlsx"))
    if not matches:
        raise FileNotFoundError("No matching MONTBELL_*.xlsx found in Downloads.")
    return matches[0]


def find_title_column(headers: list[object]) -> int:
    for idx, header in enumerate(headers, start=1):
        if isinstance(header, str) and header.strip().lower() == TITLE_COL_NAME.lower():
            return idx
    raise ValueError("Could not find the Title column.")


def ensure_adjacent_titlechn_column(ws, title_col_idx: int) -> int:
    adjacent_header = ws.cell(row=1, column=title_col_idx + 1).value
    if adjacent_header == TITLE_CHN_COL_NAME:
        return title_col_idx + 1

    existing_idx = None
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str) and value.strip().lower() == TITLE_CHN_COL_NAME.lower():
            existing_idx = col
            break

    if existing_idx is None:
        ws.insert_cols(title_col_idx + 1)
        ws.cell(row=1, column=title_col_idx + 1, value=TITLE_CHN_COL_NAME)
        return title_col_idx + 1

    ws.cell(row=1, column=existing_idx, value=None)
    ws.insert_cols(title_col_idx + 1)
    ws.cell(row=1, column=title_col_idx + 1, value=TITLE_CHN_COL_NAME)
    return title_col_idx + 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Translate Japanese product titles in an Excel file into Chinese.")
    parser.add_argument("--input", help="Source xlsx path")
    parser.add_argument("--output", help="Output xlsx path")
    parser.add_argument("--cache", default=str(CACHE_FILE), help="Translation cache json")
    args = parser.parse_args()

    input_path = resolve_input_path(args.input)
    output_path = Path(args.output) if args.output else Path.cwd() / f"{input_path.stem}_TitleChn.xlsx"
    cache_path = Path(args.cache)

    wb = load_workbook(input_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    title_col_idx = find_title_column(headers)
    titlechn_col_idx = ensure_adjacent_titlechn_column(ws, title_col_idx)

    translator = GoogleTranslator(source="auto", target="zh-CN")
    cache = load_cache(cache_path)

    unique_titles: dict[str, str] = {}
    previous_source = None
    previous_result = None

    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=title_col_idx).value
        source = "" if value is None else str(value).strip()
        if not source:
            continue
        if source == previous_source and previous_result is not None:
            unique_titles[source] = previous_result
        elif source not in unique_titles:
            unique_titles[source] = ""
        previous_source = source
        previous_result = unique_titles.get(source)

    total_unique = len(unique_titles)
    for idx, source in enumerate(unique_titles.keys(), start=1):
        if unique_titles[source]:
            continue
        unique_titles[source] = translate_title(source, translator, cache)
        if idx % 200 == 0:
            print(f"Translated unique titles {idx}/{total_unique}...", flush=True)

    previous_source = None
    previous_result = None
    total_rows = ws.max_row - 1
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=title_col_idx).value
        source = "" if value is None else str(value).strip()
        if not source:
            result = ""
        elif source == previous_source and previous_result is not None:
            result = previous_result
        else:
            result = unique_titles.get(source, source)
        ws.cell(row=row_idx, column=titlechn_col_idx, value=result)
        previous_source = source
        previous_result = result
        if (row_idx - 1) % 2000 == 0:
            print(f"Filled rows {row_idx - 1}/{total_rows}...", flush=True)

    wb.save(output_path)
    save_cache(cache_path, cache)
    print(f"Saved: {output_path}")
    print(f"Unique title translations: {len(unique_titles)}")
    print(f"Cached online translations: {len(cache)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
