from __future__ import annotations

"""patagonia.jp 在庫・価格スクレイパー / Patagonia Japan price & store-stock scraper.

items.xlsx の sku 列を読み、各 SKU について patagonia.jp を検索し、
商品詳細ページで色・サイズを選択して価格と「ストアの在庫状況」を取得し、
Excel に出力する。

SKU 形式: {item_code}-{color}-{size}  例) 60421-EDBL-18M / 29157-BLK-L-XL
- item_code: 先頭セグメント (商品番号)
- color:     2 番目のセグメント
- size:      残り全部 (ハイフンを含むことがある)

反ボット対策:
- patagonia.jp はボットを検知すると失敗用ページ (SPA-sitefailover / "Sit tight")
  を返す。Cloudflare ではないので solve_cloudflare は効かない。
- ヘッドレスは検知されやすいため headless=False (ヘッドフル) で動かし、
  最初に /home/ を開いてセッションを温める。失敗用ページを検知したら再試行する。

店舗在庫:
- 「ストアの在庫状況」は third-party の Locally で読み込まれる。色とサイズを
  選択してから a.change-store をクリックするとモーダルに店舗一覧が出る。
  そのため Scrapling の page_action で実ブラウザを操作して取得する。
"""

import argparse
import gc
import json
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from scrapling.fetchers import StealthySession


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "items.xlsx"
BASE_URL = "https://www.patagonia.jp"
HOME_URL = f"{BASE_URL}/home/"
SEARCH_URL = f"{BASE_URL}/search?q={{q}}"

STEALTH_OPTIONS = {
    "headless": False,          # ヘッドフル必須 (ヘッドレスは failover でブロックされる)
    "network_idle": True,
    "humanize": True,
    "os_randomize": True,
    "google_search": True,      # referer を google にして自然に見せる
    "disable_resources": False, # 画像/JS を全部読み込ませる (描画完了のため)
}

# セッションを温め直す間隔 (商品ページ数)。長時間連続アクセスでブロックされるのを防ぐ。
RECYCLE_EVERY = 25
# リクエスト間の待機 (秒)。礼儀的かつ反ボット回避。
DELAY_BETWEEN = 1.5
# failover ページ判定の閾値 (バイト)
FAILOVER_MIN_LEN = 30000

PRICE_RE = re.compile(r'itemprop="price"[^>]*content="([\d.,]+)"')
PRICE_RE2 = re.compile(r'class="value"[^>]*content="([\d.,]+)"')
TITLE_RE = re.compile(r"<h1[^>]*class=\"[^\"]*product-name[^\"]*\"[^>]*>(.*?)</h1>", re.I | re.S)
TITLE_FALLBACK_RE = re.compile(r"<title[^>]*>(.*?)</title>", re.I | re.S)
PRODUCT_LINK_RE_TMPL = r'/product/[^"\'\s]*?/{code}\.html'

# モーダル内の店舗在庫。card-body 単位で店舗名と在庫状態を抽出する。
# 在庫状態は Locally が非同期で埋める: dot クラス in=あり / low=わずか / out=なし
STORE_NAME_RE = re.compile(r'class="col store-name"[^>]*>\s*<a[^>]*>(.*?)</a>', re.S)
STORE_STATUS_RE = re.compile(r'class="store-availability-message"[^>]*>(.*?)<div', re.S)
STORE_DOT_RE = re.compile(r'class="dot\s+(\w+)"')
DOT_STATUS = {"in": "あり", "low": "わずか", "out": "なし"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Patagonia Japan price & store-stock scraper")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="入力 Excel (sku 列が必要)")
    parser.add_argument("--output-dir", default=str(SCRIPT_DIR), help="出力先ディレクトリ")
    parser.add_argument("--limit", type=int, default=None, help="先頭 N 件だけ処理 (テスト用)")
    parser.add_argument("--no-resume", action="store_true", help="チェックポイントを無視して最初から")
    parser.add_argument("--headful", dest="headless", action="store_false", help="(既定) ブラウザ画面を表示")
    parser.add_argument("--headless", dest="headless", action="store_true", help="ヘッドレスで実行 (ブロックされやすい)")
    parser.set_defaults(headless=False)
    return parser.parse_args()


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S")


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", unescape(re.sub(r"<[^>]+>", " ", text))).strip()


@dataclass(slots=True)
class StoreStock:
    name: str
    status: str


@dataclass(slots=True)
class ResultRow:
    sku: str
    product_name: str = ""
    url: str = ""
    color: str = ""
    color_name: str = ""
    size: str = ""
    price: str = ""
    note: str = ""
    stores: list[StoreStock] = field(default_factory=list)


def size_candidates(size: str) -> list[str]:
    """サイズ表記のゆらぎを吸収する候補リスト。L-XL -> [L-XL, L/XL] など。"""
    cands = [size, size.replace("-", "/")]
    out: list[str] = []
    for c in cands:
        if c not in out:
            out.append(c)
    return out


def parse_sku(sku: str) -> tuple[str, str, str]:
    """`60421-EDBL-18M` -> ('60421', 'EDBL', '18M'); `29157-BLK-L-XL` -> ('29157','BLK','L-XL')."""
    parts = sku.strip().split("-")
    if len(parts) < 3:
        raise ValueError(f"SKU の形式が不正です: {sku!r} (item-color-size を期待)")
    return parts[0], parts[1], "-".join(parts[2:])


def read_skus(input_path: Path, limit: int | None) -> list[str]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    try:
        sku_idx = header.index("sku")
    except ValueError as exc:
        raise ValueError(f"'sku' 列が見つかりません。ヘッダー: {header}") from exc
    skus: list[str] = []
    for row in rows:
        if sku_idx >= len(row):
            continue
        value = row[sku_idx]
        if value is None:
            continue
        sku = str(value).strip()
        if sku:
            skus.append(sku)
    wb.close()
    if limit is not None:
        skus = skus[:limit]
    return skus


def is_failover(html: str) -> bool:
    if len(html) < FAILOVER_MIN_LEN:
        return True
    lowered = html[:5000].lower()
    return "spa-sitefailover" in lowered or "botfailover" in lowered or "sit tight" in lowered


class Browser:
    """ヘッドフルな StealthySession を保持し、定期的に温め直す薄いラッパー。"""

    def __init__(self, options: dict[str, Any]) -> None:
        self.options = options
        self.session: StealthySession | None = None
        self.count = 0
        self.warmed = False

    def __enter__(self) -> "Browser":
        return self

    def __exit__(self, *exc: Any) -> None:
        self.close()

    def close(self) -> None:
        if self.session is not None:
            try:
                self.session.close()
            finally:
                self.session = None
                self.warmed = False
                gc.collect()

    def _ensure(self) -> StealthySession:
        if self.session is None:
            log("ブラウザセッションを起動中 ...")
            self.session = StealthySession(**self.options)
            self.session.start()
            self.warmed = False
        if not self.warmed:
            self._warmup()
        return self.session

    def _warmup(self) -> None:
        assert self.session is not None
        try:
            self.session.fetch(HOME_URL, network_idle=True)
            self.warmed = True
            log("セッションを温めました (/home/)")
            time.sleep(DELAY_BETWEEN)
        except Exception as exc:  # noqa: BLE001
            log(f"ウォームアップ失敗 (継続します): {exc}")
            self.warmed = True

    def recycle(self) -> None:
        log(f"{self.count} 件処理したのでセッションを再起動します")
        self.close()

    def maybe_recycle(self) -> None:
        if RECYCLE_EVERY > 0 and self.count > 0 and self.count % RECYCLE_EVERY == 0:
            self.recycle()

    def _html(self, response: Any) -> str:
        body = getattr(response, "body", b"")
        if isinstance(body, bytes) and body:
            return body.decode("utf-8", "ignore")
        text = getattr(response, "text", "")
        return text if isinstance(text, str) else ""

    def fetch(self, url: str, *, retries: int = 3, **kwargs: Any) -> str:
        """failover を検知したら待って再試行しつつ HTML を返す。"""
        last = ""
        for attempt in range(retries + 1):
            session = self._ensure()
            try:
                response = session.fetch(url, **kwargs)
            except Exception as exc:  # noqa: BLE001
                log(f"  fetch 例外 (試行 {attempt + 1}/{retries + 1}): {exc}")
                self.recycle()
                time.sleep(DELAY_BETWEEN * 2)
                continue
            html = self._html(response)
            if not is_failover(html):
                return html
            last = html
            log(f"  failover ページ検知 (試行 {attempt + 1}/{retries + 1}) len={len(html)}")
            # failover 後はセッションを作り直してから待つ
            self.recycle()
            time.sleep(DELAY_BETWEEN * (attempt + 2))
        return last

    def fetch_with_action(self, url: str, action: Callable, *, retries: int = 3, **kwargs: Any) -> str:
        """page_action 付き fetch。action 内で最終 DOM を sink に格納させて返す。"""
        last = ""
        for attempt in range(retries + 1):
            session = self._ensure()
            try:
                session.fetch(url, page_action=action, **kwargs)
            except Exception as exc:  # noqa: BLE001
                log(f"  fetch(action) 例外 (試行 {attempt + 1}/{retries + 1}): {exc}")
                self.recycle()
                time.sleep(DELAY_BETWEEN * 2)
                continue
            html = action.sink.get("html", "")  # type: ignore[attr-defined]
            if html and not is_failover(html):
                return html
            last = html
            log(f"  failover/空ページ (試行 {attempt + 1}/{retries + 1}) len={len(html)}")
            self.recycle()
            time.sleep(DELAY_BETWEEN * (attempt + 2))
        return last


def make_page_action(color: str, size: str) -> Callable:
    """色→サイズを選択し「ストアの在庫状況」を開いて最終 DOM を sink['html'] に保存する page_action。"""
    sink: dict[str, str] = {}

    def action(page: Any) -> Any:
        try:
            # swatch が出れば描画完了とみなす (networkidle を上限付きで待つ)
            try:
                page.wait_for_selector(".color-attribute, [data-attr-value]", timeout=12000)
            except Exception:  # noqa: BLE001
                try:
                    page.wait_for_load_state("networkidle", timeout=8000)
                except Exception:  # noqa: BLE001
                    pass
            page.wait_for_timeout(800)

            # 1) 色を選択 (URL で指定済みだが念のため swatch もクリック)
            for sel in (f"[data-attr-value='{color}']", f"[data-color='{color}']"):
                try:
                    page.click(sel, timeout=3000)
                    break
                except Exception:  # noqa: BLE001
                    continue
            page.wait_for_timeout(600)

            # 2) サイズを選択
            # SKU はサイズをハイフン区切りで持つが、サイト上は "/" のことがある
            # (例: SKU の L-XL = サイト上の L/XL)。候補を複数試す。
            size_clicked = False
            for sz in size_candidates(size):
                enc = quote(sz, safe="")
                size_selectors = (
                    f"label[data-size-url*='_size={enc}&']",
                    f"label[data-size-url*='_size={sz}&']",
                    f"[data-attr-value='{sz}']",
                    f"label[data-attr-value='{sz}']",
                )
                for sel in size_selectors:
                    if page.query_selector(sel) is None:
                        continue
                    try:
                        page.click(sel, timeout=3000)
                        size_clicked = True
                        break
                    except Exception:  # noqa: BLE001
                        continue
                if size_clicked:
                    break
            page.wait_for_timeout(1000)
            sink["size_clicked"] = "1" if size_clicked else "0"

            # 3)「ストアの在庫状況」を開く
            if size_clicked:
                opened = False
                for sel in ("a.change-store", ".js-popup-availability a", ".view-store-availability a"):
                    try:
                        page.click(sel, timeout=4000)
                        opened = True
                        break
                    except Exception:  # noqa: BLE001
                        continue
                if opened:
                    # まず店舗一覧の枠が出るのを待つ
                    try:
                        page.wait_for_selector(
                            ".results__stores .store-name", timeout=15000
                        )
                    except Exception:  # noqa: BLE001
                        pass
                    # 在庫状態 (dot) は Locally が非同期で埋めるので dot 出現を待つ
                    try:
                        page.wait_for_selector(
                            ".results__stores .store-availability-message .dot",
                            timeout=15000,
                        )
                    except Exception:  # noqa: BLE001
                        pass
                    page.wait_for_timeout(1200)
        finally:
            try:
                sink["html"] = page.content()
            except Exception:  # noqa: BLE001
                sink["html"] = sink.get("html", "")
        return page

    action.sink = sink  # type: ignore[attr-defined]
    return action


def find_product_url(html: str, item_code: str) -> str | None:
    pattern = re.compile(PRODUCT_LINK_RE_TMPL.format(code=re.escape(item_code)))
    match = pattern.search(html)
    if match:
        return BASE_URL + match.group(0)
    return None


def extract_price(html: str) -> str:
    for rx in (PRICE_RE, PRICE_RE2):
        match = rx.search(html)
        if match:
            return match.group(1).replace(",", "")
    return ""


def extract_product_name(html: str) -> str:
    match = TITLE_RE.search(html)
    if match:
        return clean(match.group(1))
    match = TITLE_FALLBACK_RE.search(html)
    if match:
        return clean(match.group(1)).replace(" - パタゴニア公式オンラインショップ", "")
    return ""


def extract_color_name(html: str, color: str) -> str:
    # swatch: data-attr-value="EDBL" ... data-caption="Eddy Blue"
    m = re.search(
        rf'data-attr-value="{re.escape(color)}"[^>]*?data-caption="([^"]*)"', html
    )
    if m:
        return clean(m.group(1))
    m = re.search(
        rf'data-caption="([^"]*)"[^>]*?data-attr-value="{re.escape(color)}"', html
    )
    if m:
        return clean(m.group(1))
    # JSON-LD: "color":"Eddy Blue",...,"sku":"60421-EDBL"
    m = re.search(rf'"color":"([^"]+)"[^}}]*?"sku":"[^"]*{re.escape(color)}"', html)
    if m:
        return clean(m.group(1))
    return ""


def extract_stores(html: str) -> list[StoreStock]:
    """モーダル DOM から直営店 (results__stores) の店舗名と在庫状態を抽出。"""
    stores: list[StoreStock] = []
    seen: set[str] = set()
    # store-tab (results__stores) の領域に限定し、dealer-tab 以降は切る
    start = html.find('class="results__stores"')
    region = html[start:] if start != -1 else html
    dealer = region.find('class="results__dealers"')
    if dealer != -1:
        region = region[:dealer]

    # 店舗ブロック (card-body) ごとに分割して解析する
    for block in re.split(r'<div class="card-body', region)[1:]:
        name_m = STORE_NAME_RE.search(block)
        if not name_m:
            continue
        name = clean(name_m.group(1))
        if not name or name in seen:
            continue
        # 状態テキスト (あり/わずか/なし) を優先、無ければ dot クラスから推定
        status_m = STORE_STATUS_RE.search(block)
        status = clean(status_m.group(1)) if status_m else ""
        if not status:
            dot_m = STORE_DOT_RE.search(block)
            if dot_m:
                status = DOT_STATUS.get(dot_m.group(1), dot_m.group(1))
        seen.add(name)
        stores.append(StoreStock(name=name, status=status))
    return stores


def scrape_one(browser: Browser, sku: str) -> ResultRow:
    item_code, color, size = parse_sku(sku)
    row = ResultRow(sku=sku, color=color, size=size)

    # 1) 検索して商品 URL を得る
    log(f"  検索: {item_code}")
    search_html = browser.fetch(SEARCH_URL.format(q=item_code), network_idle=True)
    browser.count += 1
    time.sleep(DELAY_BETWEEN)
    product_url = find_product_url(search_html, item_code)
    if not product_url:
        row.note = "検索で商品が見つかりません"
        log(f"  -> 見つかりません ({item_code})")
        return row
    # 色を URL で事前選択
    sep = "&" if "?" in product_url else "?"
    product_url_color = f"{product_url}{sep}dwvar_{item_code}_color={color}"
    row.url = product_url

    # 2) 商品ページを開き、色/サイズ選択 + 店舗在庫モーダルを取得
    log(f"  商品ページ: {product_url_color}")
    browser.maybe_recycle()
    action = make_page_action(color, size)
    html = browser.fetch_with_action(product_url_color, action, network_idle=True, wait=3000)
    browser.count += 1
    time.sleep(DELAY_BETWEEN)

    if not html or is_failover(html):
        row.note = "商品ページ取得失敗 (failover)"
        return row

    row.product_name = extract_product_name(html)
    row.price = extract_price(html)
    row.color_name = extract_color_name(html, color)
    if action.sink.get("size_clicked") == "0":  # type: ignore[attr-defined]
        row.note = f"サイズ {size} を選択できませんでした (在庫切れ等)"
    row.stores = extract_stores(html)
    log(f"  -> 価格={row.price or '?'} 店舗数={len(row.stores)} {row.note}")
    return row


# ---- チェックポイント ----

def checkpoint_path(output_dir: Path) -> Path:
    return output_dir / ".patagonia_cache" / "progress.json"


def load_checkpoint(path: Path, resume: bool) -> dict[str, Any]:
    if resume and path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {"results": {}}


def save_checkpoint(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def row_to_dict(row: ResultRow) -> dict[str, Any]:
    return {
        "sku": row.sku,
        "product_name": row.product_name,
        "url": row.url,
        "color": row.color,
        "color_name": row.color_name,
        "size": row.size,
        "price": row.price,
        "note": row.note,
        "stores": [{"name": s.name, "status": s.status} for s in row.stores],
    }


def dict_to_row(data: dict[str, Any]) -> ResultRow:
    row = ResultRow(
        sku=data.get("sku", ""),
        product_name=data.get("product_name", ""),
        url=data.get("url", ""),
        color=data.get("color", ""),
        color_name=data.get("color_name", ""),
        size=data.get("size", ""),
        price=data.get("price", ""),
        note=data.get("note", ""),
    )
    row.stores = [StoreStock(name=s.get("name", ""), status=s.get("status", "")) for s in data.get("stores", [])]
    return row


# ---- 出力 ----

def write_excel(rows: list[ResultRow], path: Path) -> None:
    max_stores = max((len(r.stores) for r in rows), default=0)
    headers = ["sku", "product_name", "url", "color", "color_name", "size", "price", "note"]
    for i in range(1, max_stores + 1):
        headers += [f"shop{i}", f"stock{i}"]

    wb = Workbook()
    ws = wb.active
    ws.title = "patagonia"
    ws.append(headers)
    for r in rows:
        values = [r.sku, r.product_name, r.url, r.color, r.color_name, r.size, r.price, r.note]
        for s in r.stores:
            values += [s.name, s.status]
        values += [""] * (len(headers) - len(values))
        ws.append(values)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = now_stamp()

    options = dict(STEALTH_OPTIONS)
    options["headless"] = args.headless

    skus = read_skus(input_path, args.limit)
    log(f"対象 SKU: {len(skus)} 件  (headless={options['headless']})")

    cp_path = checkpoint_path(output_dir)
    resume = not args.no_resume
    checkpoint = load_checkpoint(cp_path, resume)
    done: dict[str, Any] = checkpoint.get("results", {})

    results: list[ResultRow] = []
    try:
        with Browser(options) as browser:
            for index, sku in enumerate(skus, start=1):
                if sku in done:
                    log(f"[{index}/{len(skus)}] {sku} (キャッシュ再利用)")
                    results.append(dict_to_row(done[sku]))
                    continue
                log(f"[{index}/{len(skus)}] {sku}")
                try:
                    row = scrape_one(browser, sku)
                except Exception as exc:  # noqa: BLE001
                    log(f"  エラー: {exc}")
                    row = ResultRow(sku=sku, note=f"エラー: {exc}")
                    _ic = sku.split("-")
                    if len(_ic) >= 3:
                        row.color, row.size = _ic[1], "-".join(_ic[2:])
                results.append(row)
                done[sku] = row_to_dict(row)
                checkpoint["results"] = done
                save_checkpoint(cp_path, checkpoint)
    finally:
        out_path = output_dir / f"patagonia_output_{stamp}.xlsx"
        write_excel(results, out_path)
        log(f"出力しました: {out_path}  ({len(results)} 行)")


if __name__ == "__main__":
    main()
